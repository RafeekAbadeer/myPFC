import os
import csv
import datetime
from PyQt5.QtGui import QPainter, QColor, QFont, QFontMetrics
from PyQt5.QtPrintSupport import QPrinter
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QTableView, QHeaderView
from PyQt5.QtCore import Qt, QRect, QSize


def export_table_data(parent, table_view, default_filename=None, export_title=None):
    """
    General function to export data from a table view

    Args:
        parent: Parent widget for dialogs
        table_view: QTableView containing the data to export
        default_filename: Optional default name for the exported file
        export_title: Optional title for the export document
    """
    if not isinstance(table_view, QTableView):
        QMessageBox.warning(parent, "Export Error", "Invalid table view provided for export.")
        return

    model = table_view.model()
    if not model or model.rowCount() == 0:
        QMessageBox.information(parent, "Export Info", "No data to export.")
        return

    # If no export title is provided, try to get one from the window title
    if not export_title:
        window_title = parent.window().windowTitle()
        export_title = f"{window_title} Data"

    # Prepare default filename
    if not default_filename:
        default_filename = f"export_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"

    # Get file path from user
    file_path, selected_filter = QFileDialog.getSaveFileName(
        parent,
        "Export Data",
        os.path.expanduser(f"~/{default_filename}"),
        "CSV Files (*.csv);;Excel Files (*.xlsx);;PDF Files (*.pdf)"
    )

    if not file_path:
        return  # User canceled

    # Determine export format based on file extension
    if file_path.lower().endswith('.csv'):
        export_to_csv(parent, table_view, file_path)
    elif file_path.lower().endswith('.xlsx'):
        export_to_excel(parent, table_view, file_path)
    elif file_path.lower().endswith('.pdf'):
        export_to_pdf(parent, table_view, file_path, export_title)
    else:
        # Add default extension if none provided
        if "CSV Files" in selected_filter:
            file_path += ".csv"
            export_to_csv(parent, table_view, file_path)
        elif "Excel Files" in selected_filter:
            file_path += ".xlsx"
            export_to_excel(parent, table_view, file_path)
        elif "PDF Files" in selected_filter:
            file_path += ".pdf"
            export_to_pdf(parent, table_view, file_path, export_title)
        else:
            QMessageBox.warning(parent, "Export Error", "Unknown export format.")
            return


def export_to_csv(parent, table_view, file_path):
    """Export table data to CSV file"""
    try:
        model = table_view.model()
        rows = model.rowCount()
        columns = model.columnCount()

        with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)

            # Write headers
            headers = []
            for column in range(columns):
                header_text = model.headerData(column, Qt.Horizontal)
                headers.append(header_text if header_text else f"Column {column + 1}")
            writer.writerow(headers)

            # Write data
            for row in range(rows):
                row_data = []
                for column in range(columns):
                    index = model.index(row, column)
                    data = model.data(index)
                    row_data.append(data if data is not None else "")
                writer.writerow(row_data)

        QMessageBox.information(parent, "Export Success", f"Data exported successfully to {file_path}")
    except Exception as e:
        QMessageBox.critical(parent, "Export Error", f"Failed to export to CSV: {str(e)}")

def export_to_excel(parent, table_view, file_path):
    """Export table data to Excel file"""
    try:
        # Try to import openpyxl - will be needed for Excel export
        import openpyxl
        from openpyxl import Workbook

        model = table_view.model()
        rows = model.rowCount()
        columns = model.columnCount()

        wb = Workbook()
        ws = wb.active
        ws.title = "Exported Data"

        # Write headers
        for column in range(columns):
            header_text = model.headerData(column, Qt.Horizontal)
            cell = ws.cell(row=1, column=column + 1)
            cell.value = header_text if header_text else f"Column {column + 1}"
            # Add some styling to headers (optional)
            cell.font = openpyxl.styles.Font(bold=True)

        # Write data
        for row in range(rows):
            for column in range(columns):
                index = model.index(row, column)
                data = model.data(index)
                ws.cell(row=row + 2, column=column + 1, value=data if data is not None else "")

        # Auto-adjust column widths (optional)
        for column in range(columns):
            column_letter = openpyxl.utils.get_column_letter(column + 1)
            max_width = 0
            for row in range(rows + 1):  # +1 to include header
                cell = ws.cell(row=row + 1, column=column + 1)
                if cell.value:
                    max_width = max(max_width, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_width + 2  # Add padding

        wb.save(file_path)
        QMessageBox.information(parent, "Export Success", f"Data exported successfully to {file_path}")
    except ImportError:
        QMessageBox.critical(parent, "Export Error",
                             "Excel export requires the openpyxl package.\n"
                             "Please install it with: pip install openpyxl")
    except Exception as e:
        QMessageBox.critical(parent, "Export Error", f"Failed to export to Excel: {str(e)}")


def export_to_pdf(parent, table_view, file_path, title=None):
    """Export table data to PDF file using ReportLab"""
    try:
        # Try to import reportlab - will be needed for PDF export
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.platypus.flowables import Flowable
        from reportlab.pdfgen.canvas import Canvas

        # For page numbers
        from reportlab.platypus.doctemplate import PageTemplate, BaseDocTemplate
        from reportlab.platypus.frames import Frame

        model = table_view.model()
        rows = model.rowCount()
        columns = model.columnCount()

        # Get a more descriptive title if not provided
        if not title:
            title = "Data Export"

        # Create the full title
        full_title = title

        # Get current date and time for the report
        current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Get headers
        headers = []
        for column in range(columns):
            header_text = model.headerData(column, Qt.Horizontal)
            headers.append(header_text if header_text else f"Column {column + 1}")

        # Identify the column indices for our problematic headers
        # We'll use these indices to assign specific widths
        problem_headers = ['Category', 'Currency', 'Nature', 'Term']
        problem_indices = [i for i, h in enumerate(headers) if h in problem_headers]

        # Define which columns should allow wrapping
        wrappable_columns = ['Description', 'Classifications', 'Credit Accounts', 'Debit Accounts']

        # Get data and calculate column widths based on content
        table_data = []
        styles = getSampleStyleSheet()

        # Create a paragraph style for cells that allows wrapping
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            wordWrap='CJK',  # Use CJK wrapping for better results
            leading=12  # Space between lines
        )

        # Create a style for headers - non-wrapping
        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            wordWrap='LO',  # LO = don't wrap long words
            leading=14
        )

        # Prepare header row - avoiding Paragraph for problematic headers
        header_row = []
        for i, header in enumerate(headers):
            if i in problem_indices:
                # Use plain string for problem headers to prevent wrapping
                header_row.append(f"<b>{header}</b>")
            else:
                # Use paragraph for other headers
                header_para = Paragraph(f"<b>{header}</b>", header_style)
                header_row.append(header_para)

        table_data.append(header_row)

        # Initial column width estimates
        # Start with a base width for all columns
        base_width = 0.8 * inch

        # Calculate minimum width needed for each header
        header_widths = []
        for i, header in enumerate(headers):
            # Problem headers get extra space to ensure no wrapping
            if i in problem_indices:
                width = len(header) * 0.12 * inch + 0.15 * inch  # Additional padding
            else:
                width = len(header) * 0.1 * inch
            header_widths.append(max(width, base_width))

        # Special wider allocation for columns that need wrapping
        for i, header in enumerate(headers):
            if header in wrappable_columns:
                header_widths[i] = max(header_widths[i], 1.5 * inch)  # Give more space to wrappable columns

        # Process data rows
        data_widths = [0] * columns  # Track width needed for data
        for row in range(rows):
            row_data = []
            for column in range(columns):
                index = model.index(row, column)
                data = model.data(index)
                cell_value = data if data is not None else ""

                # Determine whether this column should have wrapping text
                is_wrappable = headers[column] in wrappable_columns

                if is_wrappable and str(cell_value).strip():
                    # Create paragraph for cell to enable wrapping
                    cell_para = Paragraph(str(cell_value).replace('\n', '<br/>'), cell_style)
                    row_data.append(cell_para)

                    # Update width estimate based on content
                    content_lines = str(cell_value).split('\n')
                    max_line_length = max(len(line) for line in content_lines) if content_lines else 0

                    # Use a reasonable estimate: each character is about 0.08-0.1 inches
                    estimated_width = max_line_length * 0.08 * inch
                    data_widths[column] = max(data_widths[column], min(estimated_width, 3.5 * inch))
                else:
                    # Regular cell text - no wrapping
                    row_data.append(str(cell_value))

                    # Update width for data cells (non-wrapping)
                    cell_width = len(str(cell_value)) * 0.1 * inch
                    data_widths[column] = max(data_widths[column], min(cell_width, 2 * inch))

            table_data.append(row_data)

        # Use the maximum of header width or data width for each column
        col_widths = [max(header_widths[i], data_widths[i]) for i in range(columns)]

        # Ensure problem headers get enough space to avoid wrapping
        for i in problem_indices:
            col_widths[i] = max(col_widths[i], len(headers[i]) * 0.15 * inch)

        # Determine if we need landscape mode by analyzing the total width
        total_estimated_width = sum(col_widths) + (0.1 * inch * columns)  # Add spacing

        # A4 sizes: 8.27 x 11.69 inches
        portrait_width = 8.27 * 0.8  # Leave margins

        # Use landscape if content is too wide for portrait
        use_landscape = total_estimated_width > portrait_width
        page_size = landscape(A4) if use_landscape else A4

        # Create a custom document template with header and footer
        class MyDocTemplate(BaseDocTemplate):
            def __init__(self, filename, **kw):
                self.allowSplitting = 0
                BaseDocTemplate.__init__(self, filename, **kw)
                template = PageTemplate('normal', [Frame(
                    self.leftMargin, self.bottomMargin, self.width, self.height, id='normal'
                )])
                self.addPageTemplates([template])

            def afterPage(self):
                self.canv.saveState()
                # Add footer with page numbers and timestamp
                self.canv.setFont('Helvetica', 8)
                footer_text = f"Page {self.canv.getPageNumber()} - Generated: {current_time}"
                self.canv.drawRightString(
                    self.width + self.leftMargin,
                    0.5 * inch,
                    footer_text
                )
                self.canv.restoreState()

        # Create the PDF document
        doc = MyDocTemplate(
            file_path,
            pagesize=page_size,
            rightMargin=0.5 * inch,
            leftMargin=0.5 * inch,
            topMargin=0.75 * inch,
            bottomMargin=0.75 * inch
        )

        # Create a list of flowables for the document
        elements = []

        # Add title
        title_style = styles['Title']
        elements.append(Paragraph(full_title, title_style))
        elements.append(Spacer(1, 0.25 * inch))

        # Check if we need to adjust column widths to fit available space
        available_width = doc.width
        total_width = sum(col_widths)

        if total_width > available_width:
            # Calculate how much we need to shrink
            scale_factor = available_width / total_width

            # Three-tiered scaling approach:
            # 1. Problem headers: least scaling
            # 2. Other non-wrappable columns: medium scaling
            # 3. Wrappable columns: most scaling
            problem_scale = min(1.0, scale_factor * 1.3)  # Try to preserve problem headers
            non_wrappable_scale = min(1.0, scale_factor * 1.1)  # Medium preservation
            wrappable_scale = scale_factor * 0.9  # Scale these more

            # Apply scaling while ensuring minimum widths
            adjusted_widths = []
            for i, width in enumerate(col_widths):
                if i in problem_indices:
                    # Problem headers get scaled the least
                    min_width = len(headers[i]) * 0.15 * inch  # Minimum to avoid wrapping
                    scaled_width = max(width * problem_scale, min_width)
                    adjusted_widths.append(scaled_width)
                elif headers[i] in wrappable_columns:
                    # Wrappable columns get scaled the most
                    adjusted_widths.append(width * wrappable_scale)
                else:
                    # Other columns get medium scaling
                    adjusted_widths.append(width * non_wrappable_scale)

            # Final check to ensure we fit
            if sum(adjusted_widths) > available_width:
                # Apply uniform scaling but preserve minimums for problem headers
                base_scaled = [w * scale_factor for w in col_widths]
                final_widths = []

                for i, width in enumerate(base_scaled):
                    if i in problem_indices:
                        min_width = len(headers[i]) * 0.12 * inch
                        final_widths.append(max(width, min_width))
                    else:
                        final_widths.append(width)

                # Last resort - if still too wide, scale everything uniformly
                if sum(final_widths) > available_width:
                    final_scale = available_width / sum(final_widths)
                    adjusted_widths = [w * final_scale for w in final_widths]
                else:
                    adjusted_widths = final_widths

            col_widths = adjusted_widths

        # Make the table with calculated widths
        table = Table(table_data, colWidths=col_widths, repeatRows=1)

        # Style the table
        style = TableStyle([
            # Header formatting
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),  # Extra padding for headers
            ('TOPPADDING', (0, 0), (-1, 0), 10),  # Extra padding for headers

            # Body formatting
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Align text to top for better wrapping

            # Default alignments
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Default left align for all

            # Table grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),

            # Zebra stripes for rows
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ])

        # Apply specific column alignments
        # 1. ID columns are usually first and should be right-aligned
        if columns > 0:
            style.add('ALIGN', (0, 1), (0, -1), 'RIGHT')

        # 2. Look for typical numeric columns like Amount, Price, etc.
        for i, header in enumerate(headers):
            if header in ['Amount', 'Price', 'Total', 'Credit Limit', 'Exchange Rate']:
                style.add('ALIGN', (i, 1), (i, -1), 'RIGHT')

        table.setStyle(style)
        elements.append(table)

        # Build the PDF
        doc.build(elements)

        QMessageBox.information(parent, "Export Success", f"Data exported successfully to {file_path}")
    except ImportError:
        QMessageBox.critical(parent, "Export Error",
                             "PDF export requires the reportlab package.\n"
                             "Please install it with: pip install reportlab")
    except Exception as e:
        QMessageBox.critical(parent, "Export Error", f"Failed to export to PDF: {str(e)}")

def is_numeric(text):
    """Check if a string represents a numeric value"""
    try:
        float(text)
        return True
    except (ValueError, TypeError):
        return False